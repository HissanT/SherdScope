import json
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from catalog.matcher_evaluation import (
    _candidate_rows,
    _candidate_values,
    export_matcher_run,
)


def _write_json(path: Path, value):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value), encoding="utf-8")


def _candidate(rank, diagnostic):
    return {
        "rank": rank,
        "reference_id": "reference-1",
        "source_filename": "reference.png",
        "citation_label": "Figure 3.1 Item 8",
        "figure": "3.1",
        "item": "8",
        "overall_score": 0.18,
        "score_components": {
            "fgw": 0.05,
            "ribbon": 0.04,
            "salience": 0.03,
            "alignment_tail": 0.02,
            "rim_region": 0.02,
            "transform_reliability": 0.02,
        },
        "fgw_cost": 0.04,
        "ribbon_cost": 0.05,
        "rim_region_cost": 0.03,
        "alignment": {
            "rms": 0.02,
            "hausdorff95": 0.04,
            "scale": 1.1,
            "rotation_degrees": 2.5,
        },
        "matched_reference_fraction": 0.44,
        "query_coverage": 1.0,
        "retrieval": {"rank": 3, "outline_rank": 4, "ribbon_rank": 2},
        "orientation_stability": 0.9,
        "warnings": [],
        "diagnostic": diagnostic,
    }


def test_evaluation_workbook_embeds_images_and_upserts_run(tmp_path):
    query_id = "a" * 32
    run_id = "b" * 32
    project = tmp_path / "project"
    query_dir = project / "matcher" / "queries" / query_id
    run_dir = project / "matcher" / "runs" / run_id
    contour_dir = project / "matcher" / "contours"
    for directory in (query_dir, run_dir, contour_dir):
        directory.mkdir(parents=True, exist_ok=True)

    for path in (
        query_dir / "query.png",
        run_dir / "diagnostic_1.png",
        run_dir / "forced.png",
        contour_dir / "reference_clean_mask.png",
    ):
        Image.new("RGB", (40, 40), "white").save(path)

    _write_json(
        query_dir / "artifact.json",
        {
            "query_id": query_id,
            "source_filename": "=query.png",
            "metadata": {
                "query_id": "HES-3.1.8",
                "rim_diameter_cm": "12.5",
                "fabric": "Fine",
                "surface": "Slip",
                "notes": "Evaluation query",
            },
        },
    )
    _write_json(
        run_dir / "result.json",
        {
            "schema_version": 1,
            "algorithm_version": "test-matcher",
            "run_id": run_id,
            "query_id": query_id,
            "confidence_margin": 0.08,
            "retrieval": {"input_count": 1099, "kept_count": 150},
            "results": [_candidate(1, "diagnostic_1.png")],
        },
    )
    _write_json(
        run_dir / "forced_result.json",
        {
            "figure": "3.1",
            "item": "8",
            "results": [_candidate("dev-1", "forced.png")],
        },
    )
    _write_json(
        contour_dir / "manifest.json",
        {
            "references": {
                "reference.png": {"clean_mask": "reference_clean_mask.png"}
            }
        },
    )

    first = export_matcher_run(project, run_id)
    second = export_matcher_run(project, run_id)

    assert first["query_count"] == 1
    assert first["result_count"] == 2
    assert first["forced_included"] is True
    assert first["already_present"] is False
    assert second["already_present"] is True
    assert second["query_count"] == 1
    assert second["result_count"] == 2

    workbook = load_workbook(first["workbook_path"])
    assert workbook.sheetnames == ["Read me", "Queries", "Ranked results"]
    assert workbook["Queries"].max_row == 2
    assert workbook["Ranked results"].max_row == 3
    assert len(workbook["Queries"]._images) == 1
    assert len(workbook["Ranked results"]._images) == 4
    assert workbook["Queries"]["C2"].value == "'=query.png"
    assert workbook["Ranked results"]["C3"].value == "Searched sherd"
    workbook.close()


def test_evaluation_keeps_shape_and_shape_plus_metadata_as_separate_rows():
    shape = _candidate(1, "diagnostic_1.png")
    fused = {
        **shape,
        "fused_rank": 1,
        "fused_score": -0.4,
        "metadata_score": 0.2,
        "metadata_weight": 0.12,
        "metadata": {
            "evidence": 0.8,
            "coverage": 0.25,
            "compared_fields": 3,
            "summary": "Metadata provides supporting evidence.",
        },
    }
    record = {
        "run_id": "run",
        "query": {"metadata": {"query_id": "Query 1"}},
        "run": {"results": [shape], "algorithm_version": "v11"},
        "metadata_result": {"results": [fused]},
    }
    rows = _candidate_rows(record)
    assert [row[0] for row in rows] == ["Top five", "Shape + metadata"]
    values = _candidate_values(record, *rows[1])
    assert len(values) == 41
    assert values[35:] == [
        -0.4,
        0.2,
        0.12,
        0.25,
        3,
        "Metadata provides supporting evidence.",
    ]
