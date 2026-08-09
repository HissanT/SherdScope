"""Persistent Excel evaluation log for partial-profile matcher runs."""

from __future__ import annotations

import json
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from catalog.contours import _atomic_json, contour_root, matcher_root, read_manifest
from catalog.matcher import load_query, load_run, query_root, run_root


WORKBOOK_FILENAME = "SherdScope_matcher_evaluation.xlsx"
RECORD_SCHEMA_VERSION = 1


class MatcherEvaluationError(ValueError):
    """Raised when a matcher run cannot be added to the evaluation workbook."""


def evaluation_root(project_path: Path) -> Path:
    return matcher_root(project_path) / "evaluation"


def evaluation_workbook_path(project_path: Path) -> Path:
    return evaluation_root(project_path) / WORKBOOK_FILENAME


def _read_json(path: Path) -> dict[str, Any] | None:
    if not path.is_file():
        return None
    try:
        with open(path, encoding="utf-8") as handle:
            value = json.load(handle)
    except (OSError, json.JSONDecodeError):
        return None
    return value if isinstance(value, dict) else None


def _record_path(project_path: Path, run_id: str) -> Path:
    return evaluation_root(project_path) / "records" / f"{run_id}.json"


def _optional_forced_result(project_path: Path, run_id: str) -> dict[str, Any] | None:
    return _read_json(run_root(project_path) / run_id / "forced_result.json")


def _optional_metadata_result(project_path: Path, run_id: str) -> dict[str, Any] | None:
    return _read_json(run_root(project_path) / run_id / "metadata_result.json")


def export_matcher_run(project_path: Path, run_id: str) -> dict[str, Any]:
    """Upsert one run record and rebuild the project evaluation workbook."""
    project_path = Path(project_path)
    run = load_run(project_path, run_id)
    query = load_query(project_path, str(run.get("query_id") or ""))
    record_path = _record_path(project_path, run_id)
    prior = _read_json(record_path)
    now = datetime.now(timezone.utc).isoformat()
    record = {
        "schema_version": RECORD_SCHEMA_VERSION,
        "run_id": run_id,
        "exported_at": (
            str(prior.get("exported_at"))
            if prior and prior.get("exported_at")
            else now
        ),
        "updated_at": now,
        "query": {
            "query_id": query.get("query_id") or run.get("query_id"),
            "source_filename": query.get("source_filename", ""),
            "metadata": query.get("metadata", {}),
        },
        "run": run,
        "forced_result": _optional_forced_result(project_path, run_id),
        "metadata_result": _optional_metadata_result(project_path, run_id),
    }
    _atomic_json(record_path, record)
    workbook_path, summary_rows, result_rows = _build_workbook(project_path)
    forced = record.get("forced_result")
    return {
        "workbook_path": workbook_path,
        "filename": workbook_path.name,
        "already_present": prior is not None,
        "query_count": summary_rows,
        "result_count": result_rows,
        "forced_included": bool(
            forced.get("results") if isinstance(forced, dict) else False
        ),
    }


def _records_from_root(root: Path) -> list[dict[str, Any]]:
    if not root.is_dir():
        return []
    records = [
        record
        for path in sorted(root.glob("*.json"))
        if (record := _read_json(path)) is not None
    ]
    return sorted(
        records,
        key=lambda item: (
            str(item.get("exported_at") or ""),
            str(item.get("run_id") or ""),
        ),
    )


def _records(project_path: Path) -> list[dict[str, Any]]:
    return _records_from_root(evaluation_root(project_path) / "records")


def _cell_value(mapping: Any, *keys: str) -> Any:
    value = mapping
    for key in keys:
        if not isinstance(value, dict):
            return None
        value = value.get(key)
    return value


def _joined(value: Any) -> str:
    if isinstance(value, (list, tuple)):
        return ", ".join(str(item) for item in value)
    return "" if value is None else str(value)


def _excel_value(value: Any) -> Any:
    """Keep user-entered labels from being interpreted as Excel formulas."""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _query_label(record: dict[str, Any]) -> str:
    metadata = _cell_value(record, "query", "metadata") or {}
    return str(
        metadata.get("query_id")
        or _cell_value(record, "query", "source_filename")
        or _cell_value(record, "query", "query_id")
        or record.get("run_id")
        or ""
    )


def _candidate_rows(record: dict[str, Any]) -> list[tuple[str, Any, dict[str, Any]]]:
    shape_results = [
        item for item in (_cell_value(record, "run", "results") or [])
        if isinstance(item, dict)
    ]
    rows = [
        ("Top five", item.get("rank"), item)
        for item in shape_results
    ]
    diagnostic_by_reference = {
        str(item.get("reference_id") or ""): item.get("diagnostic")
        for item in shape_results
    }
    metadata_result = record.get("metadata_result")
    if isinstance(metadata_result, dict):
        for item in metadata_result.get("results") or []:
            if not isinstance(item, dict):
                continue
            item = dict(item)
            item.setdefault(
                "diagnostic",
                diagnostic_by_reference.get(str(item.get("reference_id") or "")),
            )
            rows.append(("Shape + metadata", item.get("fused_rank"), item))
    forced = record.get("forced_result")
    if isinstance(forced, dict):
        rows.extend(
            ("Searched sherd", "6th / searched", item)
            for item in (forced.get("results") or [])
            if isinstance(item, dict)
        )
    return rows


def _candidate_values(
    record: dict[str, Any],
    result_kind: str,
    rank: Any,
    item: dict[str, Any],
) -> list[Any]:
    run = record.get("run") or {}
    components = item.get("score_components") or {}
    alignment = item.get("alignment") or {}
    retrieval = item.get("retrieval") or {}
    warnings = item.get("warnings") or []
    metadata = item.get("metadata") or {}
    return [
        _query_label(record),
        record.get("run_id"),
        result_kind,
        rank,
        None,
        None,
        item.get("citation_label"),
        item.get("figure"),
        item.get("item"),
        item.get("source_filename"),
        item.get("reference_id"),
        item.get("overall_score"),
        run.get("confidence_margin"),
        components.get("fgw"),
        components.get("ribbon"),
        components.get("salience"),
        components.get("alignment_tail"),
        components.get("rim_region"),
        components.get("transform_reliability"),
        components.get("completeness"),
        item.get("fgw_cost"),
        item.get("ribbon_cost", item.get("three_curve_cost")),
        item.get("rim_region_cost"),
        alignment.get("rms"),
        alignment.get("hausdorff95"),
        alignment.get("scale"),
        alignment.get("rotation_degrees"),
        item.get("matched_reference_fraction"),
        item.get("query_coverage"),
        retrieval.get("rank"),
        retrieval.get("outline_rank"),
        retrieval.get("ribbon_rank"),
        item.get("orientation_stability", item.get("initialization_stability")),
        _joined(warnings),
        run.get("algorithm_version"),
        item.get("fused_score"),
        item.get("metadata_score"),
        item.get("metadata_weight"),
        metadata.get("coverage"),
        metadata.get("compared_fields"),
        metadata.get("summary"),
    ]


def _reference_image_path(
    project_path: Path,
    manifest: dict[str, Any],
    item: dict[str, Any],
) -> Path | None:
    entry = manifest.get("references", {}).get(str(item.get("source_filename") or ""), {})
    filename = entry.get("clean_mask") or entry.get("preview")
    if not filename:
        return None
    path = contour_root(project_path) / str(filename)
    return path if path.is_file() else None


def _diagnostic_image_path(
    project_path: Path, record: dict[str, Any], item: dict[str, Any]
) -> Path | None:
    filename = str(item.get("diagnostic") or "")
    if not filename:
        return None
    path = run_root(project_path) / str(record.get("run_id") or "") / filename
    return path if path.is_file() else None


def _insert_image(sheet, path: Path | None, cell: str, *, width: int, height: int) -> None:
    if path is None:
        return
    try:
        image = WorksheetImage(str(path))
    except (OSError, ValueError):
        return
    image.width = width
    image.height = height
    sheet.add_image(image, cell)


def _style_header(sheet, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, index, header)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _set_widths(sheet, widths: dict[int, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def _add_readme(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Read me"
    rows = [
        ("SherdScope matcher evaluation workbook", "Project-level log of exported query runs."),
        ("Interpretation", "Lower match cost means a closer shape match; it is a distance, not a confidence probability."),
        ("Ranking", "Each exported run contains ranks 1–5 and the most recently scored Figure + Item, when present."),
        ("Metadata", "The workbook compares shape-only against a continuous shape-plus-metadata score. Shape has the larger weight; metadata has a smaller weight that grows smoothly with the reliable information available. A combined score can move a candidate out of the final five, but metadata cannot import a shape-rejected reference."),
        ("Images", "Query, cleaned reference, and aligned diagnostic images are embedded in the workbook."),
        ("Duplicate safety", "Exporting the same run again updates its record instead of adding duplicate rows."),
        ("Fractions", "Matched reference fraction and query coverage are stored as decimal fractions and displayed as percentages."),
    ]
    for row_index, (label, explanation) in enumerate(rows, start=1):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, explanation)
        sheet.cell(row_index, 1).font = Font(bold=True, color="1F4E78")
        sheet.cell(row_index, 1).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row_index, 2).alignment = Alignment(vertical="top", wrap_text=True)
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 105


SUMMARY_HEADERS = [
    "Query image", "Query label", "Query filename", "Prepared query UUID",
    "Run ID", "First exported (UTC)", "Diameter (cm)", "Fabric", "Surface",
    "Notes", "Reference library size", "Retrieval candidates kept",
    "First–second margin", "Searched figure", "Searched item", "Rows exported",
    "Matcher algorithm",
]

RESULT_HEADERS = [
    "Query label", "Run ID", "Result type", "Rank", "Reference image",
    "Aligned diagnostic", "Citation", "Figure", "Item", "Source filename",
    "Reference ID", "Match cost ↓", "First–second margin", "FGW component",
    "Ribbon component", "Salience component", "Alignment-tail component",
    "Rim-region component", "Transform-reliability component",
    "Completeness component", "Ordered FGW cost", "Ribbon cost",
    "Rim-region cost", "RMS residual",
    "Hausdorff 95", "Scale ratio", "Rotation (degrees)",
    "Matched reference fraction", "Query coverage", "Retrieval rank",
    "Outline retrieval rank", "Ribbon retrieval rank", "Orientation stability",
    "Warnings", "Matcher algorithm",
    "Combined score (lower is better)", "Metadata cost (lower is better)", "Metadata weight",
    "Metadata coverage", "Metadata fields compared", "Metadata explanation",
]


def _build_workbook(
    project_path: Path,
    *,
    records: list[dict[str, Any]] | None = None,
    destination: Path | None = None,
) -> tuple[Path, int, int]:
    records = _records(project_path) if records is None else records
    manifest = read_manifest(project_path)
    workbook = Workbook()
    _add_readme(workbook)

    summary = workbook.create_sheet("Queries")
    _style_header(summary, SUMMARY_HEADERS)
    _set_widths(
        summary,
        {
            1: 18, 2: 24, 3: 34, 4: 34, 5: 34, 6: 23, 7: 14, 8: 20,
            9: 20, 10: 38, 11: 18, 12: 20, 13: 18, 14: 16, 15: 14,
            16: 14, 17: 28,
        },
    )

    results_sheet = workbook.create_sheet("Ranked results")
    _style_header(results_sheet, RESULT_HEADERS)
    _set_widths(
        results_sheet,
        {
            1: 24, 2: 34, 3: 18, 4: 15, 5: 18, 6: 18, 7: 28, 8: 12,
            9: 10, 10: 38, 11: 28, 12: 16, 13: 18, 14: 16, 15: 16,
            16: 17, 17: 20, 18: 18, 19: 24, 20: 17, 21: 16, 22: 17,
            23: 16, 24: 16, 25: 14, 26: 18, 27: 22, 28: 16, 29: 15,
            30: 20, 31: 20, 32: 20, 33: 42, 34: 28, 35: 28, 36: 20,
            37: 18, 38: 18, 39: 18, 40: 20, 41: 55,
        },
    )

    result_row = 2
    for summary_row, record in enumerate(records, start=2):
        query = record.get("query") or {}
        metadata = query.get("metadata") or {}
        run = record.get("run") or {}
        retrieval = run.get("retrieval") or {}
        forced = record.get("forced_result")
        candidate_rows = _candidate_rows(record)
        summary_values = [
            None,
            _query_label(record),
            query.get("source_filename"),
            query.get("query_id"),
            record.get("run_id"),
            record.get("exported_at"),
            metadata.get("rim_diameter_cm"),
            metadata.get("fabric"),
            metadata.get("surface"),
            metadata.get("notes"),
            retrieval.get("input_count"),
            retrieval.get("kept_count"),
            run.get("confidence_margin"),
            forced.get("figure") if isinstance(forced, dict) else None,
            forced.get("item") if isinstance(forced, dict) else None,
            len(candidate_rows),
            run.get("algorithm_version"),
        ]
        for column, value in enumerate(summary_values, start=1):
            summary.cell(summary_row, column, _excel_value(value))
        summary.row_dimensions[summary_row].height = 82
        query_image = (
            query_root(project_path)
            / str(query.get("query_id") or "")
            / "query.png"
        )
        _insert_image(
            summary,
            query_image if query_image.is_file() else None,
            f"A{summary_row}",
            width=96,
            height=96,
        )

        for result_kind, rank, item in candidate_rows:
            values = _candidate_values(record, result_kind, rank, item)
            for column, value in enumerate(values, start=1):
                results_sheet.cell(result_row, column, _excel_value(value))
            results_sheet.row_dimensions[result_row].height = 82
            _insert_image(
                results_sheet,
                _reference_image_path(project_path, manifest, item),
                f"E{result_row}",
                width=96,
                height=96,
            )
            _insert_image(
                results_sheet,
                _diagnostic_image_path(project_path, record, item),
                f"F{result_row}",
                width=96,
                height=96,
            )
            for column in range(1, len(RESULT_HEADERS) + 1):
                results_sheet.cell(result_row, column).alignment = Alignment(
                    vertical="top", wrap_text=column in {7, 10, 11, 33, 34, 41}
                )
            for column in (28, 29, 38, 39):
                results_sheet.cell(result_row, column).number_format = "0.0%"
            result_row += 1

    for column in (12, 13, 14, 15, 16, 17, 18, 19, 20):
        for row in range(2, result_row):
            results_sheet.cell(row, column).number_format = "0.0000"
    for column in (36, 37):
        for row in range(2, result_row):
            results_sheet.cell(row, column).number_format = "0.0000"
    if result_row > 2:
        results_sheet.conditional_formatting.add(
            f"L2:L{result_row - 1}",
            ColorScaleRule(
                start_type="min",
                start_color="63BE7B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="F8696B",
            ),
        )
        results_sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(RESULT_HEADERS))}{result_row - 1}"
        )
    if records:
        summary.auto_filter.ref = (
            f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{len(records) + 1}"
        )

    destination = destination or evaluation_workbook_path(project_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(
        f".{destination.stem}.{uuid.uuid4().hex}.tmp{destination.suffix}"
    )
    try:
        workbook.save(temporary)
        os.replace(temporary, destination)
    except PermissionError as exc:
        raise MatcherEvaluationError(
            "Close the matcher evaluation workbook in Excel, then try again."
        ) from exc
    except OSError as exc:
        raise MatcherEvaluationError(
            f"Could not save the matcher evaluation workbook: {exc}"
        ) from exc
    finally:
        workbook.close()
        if temporary.exists():
            try:
                temporary.unlink()
            except OSError:
                pass
    return destination, len(records), result_row - 2


def export_matcher_run_to_directory(
    project_path: Path,
    run_id: str,
    destination_root: Path,
) -> dict[str, Any]:
    """Export one run into an isolated evaluation set and rebuild its workbook."""
    project_path = Path(project_path)
    destination_root = Path(destination_root)
    run = load_run(project_path, run_id)
    query = load_query(project_path, str(run.get("query_id") or ""))
    record_path = destination_root / "records" / f"{run_id}.json"
    prior = _read_json(record_path)
    now = datetime.now(timezone.utc).isoformat()
    record = {
        "schema_version": RECORD_SCHEMA_VERSION,
        "run_id": run_id,
        "exported_at": str(prior.get("exported_at")) if prior else now,
        "updated_at": now,
        "query": {
            "query_id": query.get("query_id") or run.get("query_id"),
            "source_filename": query.get("source_filename", ""),
            "metadata": query.get("metadata", {}),
        },
        "run": run,
        "forced_result": _optional_forced_result(project_path, run_id),
        "metadata_result": _optional_metadata_result(project_path, run_id),
    }
    _atomic_json(record_path, record)
    records = _records_from_root(destination_root / "records")
    workbook_path, summary_rows, result_rows = _build_workbook(
        project_path,
        records=records,
        destination=destination_root / WORKBOOK_FILENAME,
    )
    return {
        "workbook_path": workbook_path,
        "filename": workbook_path.name,
        "already_present": prior is not None,
        "query_count": summary_rows,
        "result_count": result_rows,
    }
